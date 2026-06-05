"""Excel ESIC TM import + query endpoints."""

import io
import json
import re
from datetime import date
from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, Query
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook

from app.core.database import get_db
from app.utils import clean_text, ALLOWED_DEPT

router = APIRouter(prefix="/esic", tags=["esic"])

# Parse filename: ESIC_TM_District-KIDE-202605_1 → (2026, 5, None) or 20260510 → (2026, 5, 10)
FILENAME_PATTERN = re.compile(r"(\d{6})")
DATE_PATTERN = re.compile(r"(\d{8})")


def parse_filename(fname: str) -> tuple[int, int, int | None]:
    """Extract (year, month, day) from filename. Day optional."""
    dm = DATE_PATTERN.search(fname)
    if dm:
        d = dm.group(1)  # YYYYMMDD
        return int(d[:4]), int(d[4:6]), int(d[6:8])
    m = FILENAME_PATTERN.search(fname)
    if m:
        ym = m.group(1)  # YYYYMM
        return int(ym[:4]), int(ym[4:6]), None
    raise ValueError(f"Cannot parse date from filename: {fname}")


def snapshot_date(year: int, month: int, day: int | None = None) -> date:
    """Build snapshot date. Default to 1st if day not specified."""
    return date(year, month, day or 1)


def split_docs(text_val: str | None) -> set[str]:
    """Split newline-separated document string into a set of titles."""
    if not text_val:
        return set()
    return {d.strip() for d in text_val.split("\n") if d.strip()}


def extract_docs(row) -> set[str]:
    """Extract all documents from a row (diakses + mtd)."""
    docs = set()
    if row.get("dokumen_diakses"):
        docs.update(split_docs(row["dokumen_diakses"]))
    if row.get("dokumen_mtd"):
        docs.update(split_docs(row["dokumen_mtd"]))
    return docs


@router.post("/import")
async def import_esic(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Upload Excel ESIC TM. Filter: SPL2 + STYR only."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "File must be .xlsx or .xls")

    try:
        year, month, day = parse_filename(file.filename)
    except ValueError as e:
        raise HTTPException(400, str(e))

    snap_date = snapshot_date(year, month, day)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    new_count = 0
    updated_count = 0
    skipped = 0
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        dept = clean_text(row[idx.get("Kode Dept", -1)] if idx.get("Kode Dept") is not None else "")
        if dept not in ALLOWED_DEPT:
            skipped += 1
            continue

        nrp = clean_text(row[idx.get("NRP", -1)] if idx.get("NRP") is not None else "")
        if not nrp:
            continue

        total += 1

        # Build monthly_metrics JSONB — capture all Jan-Dec columns
        monthly = {}
        for m_name in ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                       "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]:
            for suffix in ["Jumlah", "MTD", "ESIC"]:
                col = f"{m_name} {suffix}"
                if col in idx and row[idx[col]] is not None:
                    val = row[idx[col]]
                    if isinstance(val, (int, float)):
                        monthly[f"{m_name.lower()}_{suffix.lower()}"] = val

        sql = text("""
            INSERT INTO tb_esic.esic_tm
                (nrp, nama, dept, divisi, distrik, posisi, snapshot_date,
                 dokumen_diakses, dokumen_mtd, monthly_metrics)
            VALUES (:nrp, :nama, :dept, :divisi, :distrik, :posisi, :snap,
                    :doc_acc, :doc_mtd, :metrics)
            ON CONFLICT (nrp, snapshot_date) DO UPDATE SET
                nama = EXCLUDED.nama,
                dept = EXCLUDED.dept,
                divisi = EXCLUDED.divisi,
                distrik = EXCLUDED.distrik,
                posisi = EXCLUDED.posisi,
                dokumen_diakses = EXCLUDED.dokumen_diakses,
                dokumen_mtd = EXCLUDED.dokumen_mtd,
                monthly_metrics = EXCLUDED.monthly_metrics
            RETURNING (xmax = 0) AS is_new
        """)

        res = await db.execute(sql, {
            "nrp": nrp,
            "nama": clean_text(row[idx.get("Nama", -1)] if idx.get("Nama") is not None else ""),
            "dept": dept,
            "divisi": clean_text(row[idx.get("Divisi", -1)] if idx.get("Divisi") is not None else ""),
            "distrik": clean_text(row[idx.get("Distrik", -1)] if idx.get("Distrik") is not None else "KIDE"),
            "posisi": clean_text(row[idx.get("Posisi", -1)] if idx.get("Posisi") is not None else ""),
            "snap": snap_date,
            "doc_acc": clean_text(row[idx.get("Dokumen Yang Diakses", -1)] if idx.get("Dokumen Yang Diakses") is not None else ""),
            "doc_mtd": clean_text(row[idx.get("Dokumen MTD", -1)] if idx.get("Dokumen MTD") is not None else ""),
            "metrics": json.dumps(monthly),
        })
        if res.scalar():
            new_count += 1
        else:
            updated_count += 1

    await db.commit()

    return {
        "filename": file.filename,
        "snapshot_date": str(snap_date),
        "total_imported": total,
        "new_records": new_count,
        "updated_records": updated_count,
        "skipped_other_dept": skipped,
    }


@router.get("/compare")
async def compare_snapshots(
    nrp: str = Query(...),
    from_date: date = Query(...),
    to_date: date = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Bandingkan dokumen antara dua snapshot.
    Return: dokumen yang ada di to_date tapi TIDAK ada di from_date.
    """
    sql = text("""
        SELECT
            a.nrp, a.nama, a.dept,
            a.snapshot_date as from_date,
            b.snapshot_date as to_date,
            a.dokumen_diakses as from_acc,
            b.dokumen_diakses as to_acc,
            a.dokumen_mtd as from_mtd,
            b.dokumen_mtd as to_mtd
        FROM tb_esic.esic_tm a
        JOIN tb_esic.esic_tm b ON a.nrp = b.nrp
        WHERE a.nrp = :nrp
          AND a.snapshot_date = :from_date
          AND b.snapshot_date = :to_date
    """)
    res = await db.execute(sql, {"nrp": nrp, "from_date": from_date, "to_date": to_date})
    row = res.mappings().first()
    if not row:
        raise HTTPException(404, f"No data for NRP {nrp} on those dates")

    new_acc = split_docs(row["to_acc"]) - split_docs(row["from_acc"])
    new_mtd = split_docs(row["to_mtd"]) - split_docs(row["from_mtd"])

    return {
        "nrp": nrp,
        "nama": row["nama"],
        "dept": row["dept"],
        "from_date": str(from_date),
        "to_date": str(to_date),
        "new_dokumen_diakses": sorted(new_acc),
        "new_dokumen_mtd": sorted(new_mtd),
        "total_new": len(new_acc) + len(new_mtd),
    }


@router.get("/documents")
async def get_documents(
    nrp: str = Query(...),
    year: int = Query(...),
    month: int = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """Semua judul dokumen yang diakses NRP di bulan tertentu."""
    sql = text("""
        SELECT snapshot_date, dokumen_diakses, dokumen_mtd
        FROM tb_esic.esic_tm
        WHERE nrp = :nrp
          AND EXTRACT(YEAR FROM snapshot_date) = :year
          AND EXTRACT(MONTH FROM snapshot_date) = :month
        ORDER BY snapshot_date
    """)
    res = await db.execute(sql, {"nrp": nrp, "year": year, "month": month})
    rows = res.mappings().all()
    if not rows:
        raise HTTPException(404, f"No ESIC data for NRP {nrp} in {year}-{str(month).zfill(2)}")

    all_docs = set()
    snapshots = []
    for r in rows:
        docs = extract_docs(dict(r))
        all_docs.update(docs)
        snapshots.append({"date": str(r["snapshot_date"]), "count": len(docs)})

    return {
        "nrp": nrp,
        "year": year,
        "month": month,
        "total_judul": len(all_docs),
        "total_snapshots": len(rows),
        "snapshots": snapshots,
        "all_judul": sorted(all_docs),
    }


@router.get("/progress")
async def get_progress(
    nrp: str = Query(...),
    from_date: date = Query(...),
    to_date: date = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Cek progres akses dokumen NRP antara dua tanggal.
    from_date = baseline, to_date = pembanding.
    Kalau stagnan → new_count = 0, status = STAGNAN.
    """
    sql = text("""
        SELECT snapshot_date, dokumen_diakses, dokumen_mtd
        FROM tb_esic.esic_tm
        WHERE nrp = :nrp
          AND snapshot_date BETWEEN :from_date AND :to_date
        ORDER BY snapshot_date
    """)
    res = await db.execute(sql, {"nrp": nrp, "from_date": from_date, "to_date": to_date})
    rows = res.mappings().all()
    if not rows:
        raise HTTPException(404, f"No data for NRP {nrp} between {from_date} and {to_date}")

    baseline_docs = extract_docs(dict(rows[0]))
    latest_docs = extract_docs(dict(rows[-1]))
    new_docs = latest_docs - baseline_docs
    stagnant = len(new_docs) == 0

    return {
        "nrp": nrp,
        "from_date": str(rows[0]["snapshot_date"]),
        "to_date": str(rows[-1]["snapshot_date"]),
        "baseline_count": len(baseline_docs),
        "latest_count": len(latest_docs),
        "new_documents": sorted(new_docs),
        "new_count": len(new_docs),
        "stagnant": stagnant,
        "status": "STAGNAN" if stagnant else "PROGRES",
        "snapshots": [
            {"date": str(r["snapshot_date"]), "count": len(extract_docs(dict(r)))}
            for r in rows
        ],
    }


@router.get("/dept-progress")
async def get_dept_progress(
    dept: str = Query(...),
    year: int = Query(...),
    month: int = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Dashboard progres per departemen di bulan tertentu.
    Per NRP: baseline (snapshot pertama) vs snapshot terakhir.
    """
    sql = text("""
        SELECT
            e.nrp, e.nama, e.posisi,
            first_snap.snapshot_date as baseline_date,
            last_snap.snapshot_date as latest_date,
            first_snap.dokumen_diakses as baseline_diakses,
            first_snap.dokumen_mtd as baseline_mtd,
            last_snap.dokumen_diakses as latest_diakses,
            last_snap.dokumen_mtd as latest_mtd
        FROM (
            SELECT DISTINCT nrp FROM tb_esic.esic_tm
            WHERE dept = :dept
              AND EXTRACT(YEAR FROM snapshot_date) = :year
              AND EXTRACT(MONTH FROM snapshot_date) = :month
        ) unique_nrp
        JOIN LATERAL (
            SELECT snapshot_date, dokumen_diakses, dokumen_mtd
            FROM tb_esic.esic_tm
            WHERE nrp = unique_nrp.nrp
              AND EXTRACT(YEAR FROM snapshot_date) = :year
              AND EXTRACT(MONTH FROM snapshot_date) = :month
            ORDER BY snapshot_date ASC LIMIT 1
        ) first_snap ON true
        JOIN LATERAL (
            SELECT snapshot_date, dokumen_diakses, dokumen_mtd
            FROM tb_esic.esic_tm
            WHERE nrp = unique_nrp.nrp
              AND EXTRACT(YEAR FROM snapshot_date) = :year
              AND EXTRACT(MONTH FROM snapshot_date) = :month
            ORDER BY snapshot_date DESC LIMIT 1
        ) last_snap ON true
        JOIN tb_esic.esic_tm e ON e.nrp = unique_nrp.nrp
            AND e.snapshot_date = last_snap.snapshot_date
    """)
    res = await db.execute(sql, {"dept": dept, "year": year, "month": month})
    rows = res.mappings().all()

    progres = []
    stagnan_list = []
    for r in rows:
        base = split_docs(r["baseline_diakses"]) | split_docs(r["baseline_mtd"])
        latest = split_docs(r["latest_diakses"]) | split_docs(r["latest_mtd"])
        new = latest - base
        entry = {
            "nrp": r["nrp"],
            "nama": r["nama"],
            "posisi": r["posisi"],
            "baseline_date": str(r["baseline_date"]),
            "latest_date": str(r["latest_date"]),
            "new_count": len(new),
            "new_documents": sorted(new),
        }
        if len(new) > 0:
            progres.append(entry)
        else:
            stagnan_list.append(entry)

    total_karyawan = len(rows)
    return {
        "dept": dept,
        "year": year,
        "month": month,
        "total_karyawan": total_karyawan,
        "progres_count": len(progres),
        "stagnan_count": len(stagnan_list),
        "progres_persen": round(len(progres) / total_karyawan * 100, 1) if total_karyawan else 0,
        "stagnan_persen": round(len(stagnan_list) / total_karyawan * 100, 1) if total_karyawan else 0,
        "progres": progres,
        "stagnan": stagnan_list,
    }


@router.get("/by-nrp/{nrp}")
async def get_esic_by_nrp(
    nrp: str,
    db: AsyncSession = Depends(get_db),
):
    """Semua snapshot ESIC untuk satu NRP."""
    sql = text("""
        SELECT id, nrp, nama, dept, divisi, distrik, posisi,
               snapshot_date, dokumen_diakses, dokumen_mtd, monthly_metrics
        FROM tb_esic.esic_tm
        WHERE nrp = :nrp
        ORDER BY snapshot_date DESC
    """)
    res = await db.execute(sql, {"nrp": nrp})
    rows = res.mappings().all()
    if not rows:
        raise HTTPException(404, f"NRP {nrp} not found")
    return [dict(r) for r in rows]


@router.get("/")
async def list_esic(
    nrp: str | None = Query(None),
    dept: str | None = Query(None),
    distrik: str | None = Query(None),
    snapshot_date: date | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
):
    """List ESIC records dengan filter."""
    filters = []
    params: dict = {}
    if nrp:
        filters.append("nrp = :nrp")
        params["nrp"] = nrp
    if dept:
        filters.append("dept = :dept")
        params["dept"] = dept
    if distrik:
        filters.append("distrik = :distrik")
        params["distrik"] = distrik
    if snapshot_date:
        filters.append("snapshot_date = :snap")
        params["snap"] = snapshot_date

    where = " AND ".join(filters) if filters else "TRUE"
    offset = (page - 1) * page_size
    params["offset"] = offset
    params["limit"] = page_size

    count_sql = text(f"SELECT COUNT(*) FROM tb_esic.esic_tm WHERE {where}")
    data_sql = text(f"""
        SELECT id, nrp, nama, dept, divisi, distrik, posisi,
               snapshot_date, dokumen_diakses, dokumen_mtd, monthly_metrics
        FROM tb_esic.esic_tm
        WHERE {where}
        ORDER BY snapshot_date DESC, nrp
        OFFSET :offset LIMIT :limit
    """)

    total = (await db.execute(count_sql, params)).scalar() or 0
    rows = (await db.execute(data_sql, params)).mappings().all()

    return {"total": total, "page": page, "page_size": page_size, "data": [dict(r) for r in rows]}
