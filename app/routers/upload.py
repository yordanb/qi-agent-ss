"""Upload endpoint — PIN dulu, baru upload."""

import io
import random
import string
from datetime import date, datetime, timedelta, timezone

from fastapi import APIRouter, File, UploadFile, HTTPException, Form
from fastapi.responses import HTMLResponse
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import Depends

from app.core.database import engine, get_db
from app.core.security import get_current_user

router = APIRouter(tags=["upload"])

ALLOWED_DEPT = {"SPL2", "STYR"}

# ── PIN via Database (independent engine.begin() — no DI session) ───────

async def refresh_pin_db(db=None):
    """Generate new PIN and store in database via raw connection."""
    pin = "".join(random.choices(string.digits, k=6))
    now = datetime.now(timezone.utc)
    expires_at = now + timedelta(seconds=120)
    async with engine.begin() as conn:
        await conn.execute(text("""
            UPDATE tb_ss.pin_state
            SET pin = :pin, expires_at = :expires_at, fail_count = 0, locked_until = NULL
            WHERE id = 1
        """), {"pin": pin, "expires_at": expires_at})
    return pin, 120

async def get_pin_info_db(db=None):
    """Return (pin, remaining_seconds) from database via raw connection."""
    async with engine.begin() as conn:
        result = await conn.execute(text("SELECT pin, expires_at FROM tb_ss.pin_state WHERE id = 1"))
        row = result.mappings().first()
    now = datetime.now(timezone.utc)
    if row and row["expires_at"] and row["pin"]:
        expires = row["expires_at"]
        if not expires.tzinfo:
            expires = expires.replace(tzinfo=timezone.utc)
        if now < expires:
            remaining = max(int((expires - now).total_seconds()), 0)
            return row["pin"], remaining
    # Expired or empty — generate new
    return await refresh_pin_db()

async def validate_pin_db(pin: str, db=None):
    """Verify PIN against database via raw connection."""
    now = datetime.now(timezone.utc)

    # Check lockout
    async with engine.begin() as conn:
        result = await conn.execute(text("SELECT fail_count, locked_until FROM tb_ss.pin_state WHERE id = 1"))
        row = result.mappings().first()
    if row and row["locked_until"]:
        lockout_until = row["locked_until"]
        if not lockout_until.tzinfo:
            lockout_until = lockout_until.replace(tzinfo=timezone.utc)
        if now < lockout_until:
            remaining = int((lockout_until - now).total_seconds())
            return {"valid": False, "detail": "locked", "remaining": remaining}
        # Lockout expired — reset
        async with engine.begin() as conn:
            await conn.execute(text("UPDATE tb_ss.pin_state SET fail_count = 0, locked_until = NULL WHERE id = 1"))

    # Read PIN directly from DB
    async with engine.begin() as conn:
        result = await conn.execute(text("SELECT pin, expires_at FROM tb_ss.pin_state WHERE id = 1"))
        pin_row = result.mappings().first()
    if not pin_row or not pin_row["pin"] or not pin_row["expires_at"]:
        return {"valid": False, "detail": "wrong", "attempts_left": 3}

    expires = pin_row["expires_at"]
    if not expires.tzinfo:
        expires = expires.replace(tzinfo=timezone.utc)
    if now >= expires:
        return {"valid": False, "detail": "wrong", "attempts_left": 3}

    current_pin = pin_row["pin"]

    if pin == current_pin:
        # Success — reset fail count
        async with engine.begin() as conn:
            await conn.execute(text("UPDATE tb_ss.pin_state SET fail_count = 0, locked_until = NULL WHERE id = 1"))
        return {"valid": True}
    else:
        # Wrong — increment fail count
        async with engine.begin() as conn:
            await conn.execute(text("UPDATE tb_ss.pin_state SET fail_count = fail_count + 1 WHERE id = 1"))

        # Re-read updated fail_count
        async with engine.begin() as conn:
            result = await conn.execute(text("SELECT fail_count FROM tb_ss.pin_state WHERE id = 1"))
            fc = result.scalar() or 0

        if fc >= 3:
            lockout_until = now + timedelta(minutes=1)
            async with engine.begin() as conn:
                await conn.execute(text("UPDATE tb_ss.pin_state SET locked_until = :lt WHERE id = 1"), {"lt": lockout_until})
            return {"valid": False, "detail": "locked", "remaining": 60}
        return {"valid": False, "detail": "wrong", "attempts_left": 3 - fc}

# ── Helpers ──────────────────────────────────────────────

def clean_text(val):
    if val is None: return ""
    s = str(val).strip().replace("\t", " ").replace("\n", " ").replace("\r", " ")
    return " ".join(s.split())

def parse_date(val):
    if val is None or str(val).strip() == "": return None
    s = str(val).strip()
    import re
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        d, mo, y = m.groups()
        try: return date(int(y), int(mo), int(d))
        except: return None
    try: return datetime.strptime(s, "%Y-%m-%d").date()
    except: return None

def parse_num(val):
    if val is None or str(val).strip() == "": return None
    import re
    n = re.sub(r"[^0-9.\-]", "", str(val))
    if not n: return None
    try: return float(n)
    except: return None

# ── Web UI ───────────────────────────────────────────────

@router.get("/update", response_class=HTMLResponse)
async def upload_page():
    return """<!DOCTYPE html>
<html lang="id">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Quality Innovation — Update Data</title>
  <style>
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body { font-family: 'Segoe UI', sans-serif; background: #f5f5f5; padding: 24px; }
    .header { text-align: center; margin-bottom: 32px; }
    .header h1 { color: #37474f; font-size: 24px; }
    .header p { color: #78909c; margin-top: 4px; }
    .container { max-width: 600px; margin: 0 auto; }

    /* Modal Overlay */
    .modal-overlay {
      display: none; position: fixed; top: 0; left: 0; right: 0; bottom: 0;
      background: rgba(0,0,0,0.5); z-index: 100;
      align-items: center; justify-content: center;
    }
    .modal-overlay.show { display: flex; }
    .modal {
      background: white; border-radius: 20px; padding: 32px;
      width: 90%; max-width: 400px; text-align: center;
      box-shadow: 0 20px 60px rgba(0,0,0,0.3);
    }
    .modal h3 { color: #37474f; margin-bottom: 8px; font-size: 20px; }
    .modal p { color: #78909c; font-size: 13px; margin-bottom: 20px; }
    .pin-digits { display: flex; flex-direction: column; gap: 10px; margin: 20px 0; }
    .pin-row { display: flex; gap: 10px; justify-content: center; }
    .pin-box {
      width: 52px; height: 60px; font-size: 28px; font-weight: bold;
      text-align: center; border: 2px solid #cfd8dc; border-radius: 10px;
      outline: none; background: #f5f5f5;
    }
    .pin-box:focus { border-color: #37474f; background: white; }
    .modal-timer { font-size: 13px; color: #78909c; margin-top: 12px; }
    .modal-timer.warning { color: #e65100; font-weight: 600; }
    .modal-msg { font-size: 12px; margin-top: 8px; min-height: 18px; }
    .modal-msg.error { color: #c62828; }
    .modal-msg.success { color: #2e7d32; }

    .upload-zone {
      border: 3px dashed #90a4ae; border-radius: 16px;
      padding: 48px 24px; text-align: center;
      background: white; cursor: pointer; transition: all 0.2s;
    }
    .upload-zone:hover, .upload-zone.dragover {
      border-color: #37474f; background: #eceff1;
    }
    .upload-zone .icon { font-size: 48px; margin-bottom: 12px; }
    .upload-zone h3 { color: #37474f; margin-bottom: 8px; }
    .upload-zone p { color: #78909c; font-size: 14px; }
    .upload-zone input { display: none; }
    .type-select { display: flex; gap: 12px; margin: 24px 0; justify-content: center; }
    .type-btn {
      padding: 12px 24px; border: 2px solid #cfd8dc;
      border-radius: 12px; background: white; cursor: pointer;
      font-size: 14px; font-weight: 600; transition: all 0.2s;
    }
    .type-btn:hover { border-color: #37474f; }
    .type-btn.selected-closed { border-color: #2e7d32; background: #e8f5e9; color: #2e7d32; }
    .type-btn.selected-outstanding { border-color: #e65100; background: #fff3e0; color: #e65100; }
    .btn {
      display: block; width: 100%; padding: 14px;
      background: #37474f; color: white; border: none;
      border-radius: 12px; font-size: 16px; font-weight: 600;
      cursor: pointer; margin-top: 16px; transition: all 0.2s;
    }
    .btn:hover { background: #263238; }
    .btn:disabled { background: #cfd8dc; cursor: not-allowed; }

    .result {
      margin-top: 24px; padding: 20px;
      background: white; border-radius: 12px; display: none;
    }
    .result.show { display: block; }
    .result h3 { margin-bottom: 12px; }
    .result .success { color: #2e7d32; }
    .result .error { color: #c62828; }
    .result table { width: 100%; border-collapse: collapse; margin-top: 8px; }
    .result td, .result th { padding: 6px 8px; text-align: left; border-bottom: 1px solid #eee; }
    .result th { color: #546e7a; font-size: 12px; }

    .spinner { display: none; text-align: center; margin-top: 16px; }
    .spinner.show { display: block; }
    .spinner::after {
      content: ''; display: inline-block; width: 32px; height: 32px;
      border: 4px solid #cfd8dc; border-top-color: #37474f;
      border-radius: 50%; animation: spin 0.8s linear infinite;
    }
    @keyframes spin { to { transform: rotate(360deg); } }

    .locked { text-align: center; padding: 40px 20px; color: #b0bec5; }
    .locked .lock-icon { font-size: 64px; margin-bottom: 16px; }
    .locked h3 { color: #78909c; margin-bottom: 8px; }
  </style>
</head>
<body>
  <div class="header" id="header" style="display:none;">
    <h1>📊 Quality Innovation — Update Data</h1>
    <p>Upload file Excel (.xlsx) untuk update data Suggestion System</p>
  </div>
  <div class="container">

    <!-- PIN Modal -->
    <div class="modal-overlay show" id="pinModal">
      <div class="modal">
        <h3>🔑 Masukkan PIN</h3>
        <div class="pin-digits" id="pinDigits">
          <div class="pin-row">
            <input type="text" class="pin-box" id="p0" maxlength="1" autocomplete="off">
            <input type="text" class="pin-box" id="p1" maxlength="1" autocomplete="off">
            <input type="text" class="pin-box" id="p2" maxlength="1" autocomplete="off">
          </div>
          <div class="pin-row">
            <input type="text" class="pin-box" id="p3" maxlength="1" autocomplete="off">
            <input type="text" class="pin-box" id="p4" maxlength="1" autocomplete="off">
            <input type="text" class="pin-box" id="p5" maxlength="1" autocomplete="off">
          </div>
        </div>
        <div id="lockoutOverlay" style="display:none; text-align:center; padding:10px 0;">
          <div style="font-size:48px; margin-bottom:8px;">🚫</div>
          <div style="font-size:14px; font-weight:600; color:#c62828;">Terlalu Banyak Percobaan</div>
          <div style="font-size:13px; color:#78909c; margin-top:4px;">Tunggu <span id="lockoutTimer">60</span> detik</div>
        </div>
        <div class="modal-timer" id="modalTimer">⏱️ Berlaku: 2:00</div>
        <div class="modal-msg" id="pinMsg"></div>
      </div>
    </div>

    <!-- Upload area (hidden until PIN verified) -->
    <div id="uploadForm" style="display:none;">
      <div class="type-select">
        <div class="type-btn selected-closed" data-type="closed" onclick="selectType('closed')">✅ Approved</div>
        <div class="type-btn" data-type="outstanding" onclick="selectType('outstanding')">⏳ Outstanding</div>
      </div>
      <div class="upload-zone" id="dropZone" onclick="document.getElementById('fileInput').click()">
        <input type="file" id="fileInput" accept=".xlsx,.xls" onchange="handleFile(this)">
        <div class="icon">📁</div>
        <h3>Seret file Excel ke sini</h3>
        <p>atau klik untuk memilih file</p>
        <p style="margin-top:8px;font-size:12px;color:#b0bec5" id="selectedFile"></p>
      </div>
      <button class="btn" id="uploadBtn" disabled onclick="submitUpload()">📤 Upload & Import</button>
    </div>

    <div class="result" id="result"></div>
    <div class="spinner" id="spinner"></div>
  </div>

  <script>
    let selectedType = 'closed';
    let selectedFile = null;
    let pinTimerInterval = null;

    const pinInput = document.getElementById('pinInput');
    const pinModal = document.getElementById('pinModal');
    const uploadForm = document.getElementById('uploadForm');
    const pinMsg = document.getElementById('pinMsg');
    const dropZone = document.getElementById('dropZone');
    const fileInput = document.getElementById('fileInput');
    const uploadBtn = document.getElementById('uploadBtn');

    // Auto-focus first box
    document.getElementById('p0').focus();

    // PIN boxes: auto-advance + auto-verify
    for (let i = 0; i < 6; i++) {
      const box = document.getElementById('p' + i);

      box.addEventListener('input', () => {
        pinMsg.textContent = '';
        pinMsg.className = 'modal-msg';
        // Only allow digits
        box.value = box.value.replace(/[^0-9]/g, '');
        // Auto-advance
        if (box.value.length === 1 && i < 5) {
          document.getElementById('p' + (i + 1)).focus();
        }
        // Auto-verify when all 6 filled
        if (i === 5 && box.value.length === 1) {
          let pin = '';
          for (let j = 0; j < 6; j++) pin += document.getElementById('p' + j).value;
          if (pin.length === 6) verifyPin(pin);
        }
      });

      box.addEventListener('keydown', (e) => {
        // Backspace on empty → go back
        if (e.key === 'Backspace' && box.value === '' && i > 0) {
          document.getElementById('p' + (i - 1)).focus();
        }
        // Enter → verify
        if (e.key === 'Enter') {
          let pin = '';
          for (let j = 0; j < 6; j++) pin += document.getElementById('p' + j).value;
          if (pin.length === 6) verifyPin(pin);
        }
      });

      // Paste support
      box.addEventListener('paste', (e) => {
        e.preventDefault();
        const pasted = (e.clipboardData || window.clipboardData).getData('text').replace(/[^0-9]/g, '').slice(0, 6);
        for (let j = 0; j < pasted.length; j++) {
          document.getElementById('p' + j).value = pasted[j];
        }
        if (pasted.length === 6) {
          verifyPin(pasted);
        } else if (pasted.length > 0) {
          document.getElementById('p' + Math.min(pasted.length, 5)).focus();
        }
      });
    }

    let lockoutInterval = null;

    function showLockout(remaining) {
      document.getElementById('pinDigits').style.display = 'none';
      document.getElementById('modalTimer').style.display = 'none';
      document.getElementById('lockoutOverlay').style.display = 'block';
      document.getElementById('lockoutTimer').textContent = remaining;

      let timeLeft = remaining;
      clearInterval(lockoutInterval);
      lockoutInterval = setInterval(() => {
        timeLeft--;
        document.getElementById('lockoutTimer').textContent = timeLeft;
        if (timeLeft <= 0) {
          clearInterval(lockoutInterval);
          hideLockout();
        }
      }, 1000);
    }

    function hideLockout() {
      document.getElementById('pinDigits').style.display = 'flex';
      document.getElementById('modalTimer').style.display = 'block';
      document.getElementById('lockoutOverlay').style.display = 'none';
      pinMsg.textContent = '';
      pinMsg.className = 'modal-msg';
      for (let j = 0; j < 6; j++) document.getElementById('p' + j).value = '';
      document.getElementById('p0').focus();
    }

    async function verifyPin(pin) {
      try {
        const resp = await fetch('/api/upload/verify-pin?pin=' + pin);
        const data = await resp.json();

        if (data.valid) {
          clearInterval(pinTimerInterval);
          clearInterval(lockoutInterval);
          pinModal.classList.remove('show');
          pinModal.style.display = 'none';
          document.getElementById('header').style.display = 'block';
          uploadForm.style.display = 'block';
        } else if (data.detail === 'locked') {
          showLockout(data.remaining || 60);
        } else {
          const left = data.attempts_left || 0;
          pinMsg.textContent = 'PIN salah! ' + (left > 0 ? left + ' kesempatan lagi' : '');
          pinMsg.className = 'modal-msg error';
          for (let j = 0; j < 6; j++) document.getElementById('p' + j).value = '';
          document.getElementById('p0').focus();
        }
      } catch (err) {
        pinMsg.textContent = 'Error: ' + err.message;
        pinMsg.className = 'modal-msg error';
      }
    }

    dropZone.addEventListener('dragover', (e) => { e.preventDefault(); dropZone.classList.add('dragover'); });
    dropZone.addEventListener('dragleave', () => dropZone.classList.remove('dragover'));
    dropZone.addEventListener('drop', (e) => {
      e.preventDefault(); dropZone.classList.remove('dragover');
      if (e.dataTransfer.files.length > 0) { fileInput.files = e.dataTransfer.files; handleFile(fileInput); }
    });

    function selectType(type) {
      selectedType = type;
      document.querySelectorAll('.type-btn').forEach(b => b.className = 'type-btn');
      document.querySelector('[data-type="' + type + '"]').classList.add('selected-' + type);
    }

    function handleFile(input) {
      const file = input.files[0];
      if (file) {
        selectedFile = file;
        document.getElementById('selectedFile').textContent = file.name + ' (' + (file.size/1024/1024).toFixed(2) + ' MB)';
        uploadBtn.disabled = false;
      }
    }

    async function submitUpload() {
      if (!selectedFile) return;
      uploadBtn.disabled = true;
      document.getElementById('spinner').classList.add('show');

      let pin = '';
      for (let j = 0; j < 6; j++) pin += document.getElementById('p' + j).value;
      const formData = new FormData();
      formData.append('file', selectedFile);
      formData.append('import_type', selectedType);
      formData.append('pin', pin);

      try {
        const resp = await fetch('/api/upload/import', { method: 'POST', body: formData });
        const data = await resp.json();
        document.getElementById('spinner').classList.remove('show');

        if (resp.ok && data.status === 'success') {
          showResult(true, data.result);
        } else {
          showResult(false, data.detail || 'Import gagal');
        }
      } catch (err) {
        document.getElementById('spinner').classList.remove('show');
        showResult(false, 'Error: ' + err.message);
      }
    }

    function showResult(success, data) {
      document.getElementById('uploadForm').style.display = 'none';
      const result = document.getElementById('result');
      result.classList.add('show');

      if (success && typeof data === 'object') {
        result.innerHTML = '<h3 class="success">✅ Import Berhasil!</h3>' +
          '<table>' +
          '<tr><th>File</th><td>' + (data.filename || '-') + '</td></tr>' +
          '<tr><th>Tipe</th><td>' + (data.source_type || '-') + '</td></tr>' +
          '<tr><th>Total Rows</th><td>' + (data.total_rows || 0).toLocaleString() + '</td></tr>' +
          '<tr><th>Filtered</th><td>' + (data.filtered_rows || 0).toLocaleString() + '</td></tr>' +
          '<tr><th>New</th><td>' + (data.new_records || 0).toLocaleString() + '</td></tr>' +
          '<tr><th>Updated</th><td>' + (data.updated_records || 0).toLocaleString() + '</td></tr>' +
          '</table>' +
          '<button class="btn" style="margin-top:16px" onclick="location.reload()">Upload Lagi</button>';
      } else {
        result.innerHTML = '<h3 class="error">❌ ' + (data || 'Error') + '</h3>' +
          '<button class="btn" style="margin-top:16px" onclick="location.reload()">Coba Lagi</button>';
      }
    }

    // PIN countdown 2 minutes
    let timeLeft = 120;
    pinTimerInterval = setInterval(() => {
      timeLeft--;
      const m = Math.floor(timeLeft / 60);
      const s = timeLeft % 60;
      const timer = document.getElementById('modalTimer');
      timer.textContent = '⏱️ Berlaku: ' + m + ':' + s.toString().padStart(2, '0');
      if (timeLeft <= 30) timer.className = 'modal-timer warning';
      if (timeLeft <= 0) {
        timeLeft = 120;
        pinInput.value = '';
        pinMsg.textContent = 'PIN expired, masukkan PIN baru';
        pinMsg.className = 'modal-msg error';
        pinInput.focus();
      }
    }, 1000);
  </script>
</body>
</html>"""

# ── API Endpoints ────────────────────────────────────────

@router.get("/upload/pin")
async def get_current_pin_api():
    """App polls this to show current PIN."""
    pin, remaining = await get_pin_info_db()
    return {"pin": pin, "remaining_seconds": remaining}

@router.get("/upload/verify-pin")
async def verify_pin(pin: str):
    """Verify PIN before allowing upload."""
    return await validate_pin_db(pin)

@router.post("/upload/import")
async def upload_and_import(
    file: UploadFile = File(...),
    import_type: str = Form("closed"),
    pin: str = Form(...),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Upload file + import langsung setelah PIN verified."""
    result = await validate_pin_db(pin)
    if not result["valid"]:
        if result["detail"] == "locked":
            raise HTTPException(423, detail=f"Terlalu banyak percobaan. Coba lagi dalam {result['remaining']}s")
        raise HTTPException(401, detail="PIN salah atau expired")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, detail="File harus .xlsx atau .xls")

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(400, detail="File kosong")

    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb["Report"] if "Report" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(400, detail="Sheet kosong")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, detail=f"File Excel tidak valid: {e}")

    session = {
        "file_content": content,
        "filename": file.filename,
        "import_type": import_type,
        "total_rows": len(rows) - 1,
        "status": "approved",
    }

    try:
        result = await run_import(db, session)
        return {"status": "success", "result": result}
    except Exception as e:
        raise HTTPException(500, detail=f"Import error: {e}")

# ── Import Engine ────────────────────────────────────────

async def run_import(db: AsyncSession, session: dict):
    content = session["file_content"]
    import_type = session["import_type"]
    filename = session["filename"]

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb["Report"] if "Report" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip().upper() if h else "" for h in rows[0]]
    data_rows = rows[1:]

    log_res = await db.execute(
        text("INSERT INTO tb_ss.import_log (filename, source_type) VALUES (:fn, :st) RETURNING id"),
        {"fn": filename, "st": import_type},
    )
    import_id = log_res.scalar_one()

    if import_type in ("closed", "both"):
        new_count, updated_count, filtered = await import_closed(db, headers, data_rows, import_id)
    else:
        new_count, updated_count, filtered = await import_outstanding(db, headers, data_rows, import_id)

    await db.execute(
        text("UPDATE tb_ss.import_log SET total_rows=:total, new_records=:new, updated_records=:upd WHERE id=:id"),
        {"total": filtered, "new": new_count, "upd": updated_count, "id": import_id},
    )
    await db.commit()

    return {
        "filename": filename,
        "source_type": import_type,
        "total_rows": len(data_rows),
        "filtered_rows": filtered,
        "new_records": new_count,
        "updated_records": updated_count,
        "import_id": import_id,
    }

async def import_closed(db, headers, data_rows, import_id):
    new_count = updated_count = filtered = 0
    for row in data_rows:
        r = dict(zip(headers, row))
        dept = clean_text(r.get("DEPT", ""))
        if dept not in ALLOWED_DEPT: continue
        filtered += 1
        no_ss = clean_text(r.get("NO SS", ""))
        if not no_ss: continue

        vals = {
            "no_ss": no_ss, "judul": clean_text(r.get("JUDUL", "")),
            "nrp": clean_text(r.get("NRP", "")), "nama": clean_text(r.get("PENCETUS IDE", "")),
            "dept": dept, "divisi": "",
            "distrik": clean_text(r.get("DISTRIK", "")) or "KIDE",
            "tanggal_laporan": parse_date(r.get("TANGGAL LAPORAN")),
            "grade_ss": clean_text(r.get("GRADE SS", "")),
            "kualitas_ss": clean_text(r.get("KUALITAS SS", "")),
            "kategori_ss": clean_text(r.get("KATEGORI SS", "")),
            "reward_ss": parse_num(r.get("REWARD SS")) or 0,
            "status_karyawan": clean_text(r.get("STATUS KARYAWAN", "")),
            "manfaat_financial": parse_num(r.get("MANFAAT FINANCIAL")),
            "tanggal_menilai": parse_date(r.get("TANGGAL MENILAI")),
            "dinilai_oleh": clean_text(r.get("DINILAI OLEH", "")),
            "current_status": "closed", "source": "closed", "import_id": import_id,
        }
        q = """INSERT INTO tb_ss.ss_records
            (no_ss,judul,nrp,nama,dept,divisi,distrik,tanggal_laporan,
             grade_ss,kualitas_ss,kategori_ss,reward_ss,status_karyawan,
             manfaat_financial,tanggal_menilai,dinilai_oleh,
             current_status,source,latest_import_id)
            VALUES (:no_ss,:judul,:nrp,:nama,:dept,:divisi,:distrik,:tanggal_laporan,
                    :grade_ss,:kualitas_ss,:kategori_ss,:reward_ss,:status_karyawan,
                    :manfaat_financial,:tanggal_menilai,:dinilai_oleh,
                    :current_status,:source,:import_id)
            ON CONFLICT (no_ss) DO UPDATE SET
              judul=EXCLUDED.judul,nrp=EXCLUDED.nrp,nama=EXCLUDED.nama,
              dept=EXCLUDED.dept,divisi=EXCLUDED.divisi,distrik=EXCLUDED.distrik,
              tanggal_laporan=EXCLUDED.tanggal_laporan,grade_ss=EXCLUDED.grade_ss,
              kualitas_ss=EXCLUDED.kualitas_ss,kategori_ss=EXCLUDED.kategori_ss,
              reward_ss=EXCLUDED.reward_ss,status_karyawan=EXCLUDED.status_karyawan,
              manfaat_financial=EXCLUDED.manfaat_financial,
              tanggal_menilai=EXCLUDED.tanggal_menilai,
              dinilai_oleh=EXCLUDED.dinilai_oleh,
              current_status=EXCLUDED.current_status,
              source=EXCLUDED.source,latest_import_id=EXCLUDED.latest_import_id
            RETURNING (xmax=0) AS is_new"""
        res = await db.execute(text(q), vals)
        if res.mappings().first()["is_new"]: new_count += 1
        else: updated_count += 1
    return new_count, updated_count, filtered

async def import_outstanding(db, headers, data_rows, import_id):
    new_count = updated_count = filtered = 0
    for row in data_rows:
        r = dict(zip(headers, row))
        dept = clean_text(r.get("DEPT", ""))
        if dept not in ALLOWED_DEPT: continue
        filtered += 1
        no_ss = clean_text(r.get("NO SS", ""))
        if not no_ss: continue

        vals = {
            "no_ss": no_ss, "judul": clean_text(r.get("JUDUL", "")),
            "nrp": clean_text(r.get("NRP", "")), "nama": clean_text(r.get("NAMA", "")),
            "dept": dept, "divisi": clean_text(r.get("DIVISI", "")),
            "distrik": clean_text(r.get("DISTRIK", "")) or "KIDE",
            "tanggal_laporan": parse_date(r.get("TANGGAL LAPORAN")),
            "current_status": clean_text(r.get("STATUS", "")) or "Pending",
            "import_id": import_id,
        }
        q = """INSERT INTO tb_ss.ss_records
            (no_ss,judul,nrp,nama,dept,divisi,distrik,tanggal_laporan,
             grade_ss,kualitas_ss,kategori_ss,reward_ss,status_karyawan,
             manfaat_financial,tanggal_menilai,dinilai_oleh,
             current_status,source,latest_import_id)
            VALUES (:no_ss,:judul,:nrp,:nama,:dept,:divisi,:distrik,:tanggal_laporan,
                    '','','',0,'',NULL,NULL,'',
                    :current_status,'outstanding',:import_id)
            ON CONFLICT (no_ss) DO UPDATE SET
              judul=EXCLUDED.judul,nrp=EXCLUDED.nrp,nama=EXCLUDED.nama,
              dept=EXCLUDED.dept,divisi=EXCLUDED.divisi,distrik=EXCLUDED.distrik,
              tanggal_laporan=EXCLUDED.tanggal_laporan,
              current_status=EXCLUDED.current_status,
              source='outstanding',latest_import_id=EXCLUDED.latest_import_id
            RETURNING (xmax=0) AS is_new"""
        res = await db.execute(text(q), vals)
        if res.mappings().first()["is_new"]: new_count += 1
        else: updated_count += 1
    return new_count, updated_count, filtered
