"""Excel import endpoints."""

import io
import uuid
from datetime import date
from fastapi import APIRouter, Depends, File, UploadFile, HTTPException
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook

from pydantic import BaseModel
from app.utils import clean_text, parse_date, parse_num, ALLOWED_DEPT


class ImportResult(BaseModel):
    success: bool
    message: str
    total: int = 0
    new: int = 0
    updated: int = 0


router = APIRouter(prefix="/import", tags=["import"])


async def _process_closed_records(db: AsyncSession, ws, import_id: int):
    """Process closed SS worksheet."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Map header -> index
    idx = {h: i for i, h in enumerate(headers)}

    new_count = 0
    updated_count = 0
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        dept = clean_text(row[idx.get("DEPT", -1)] if idx.get("DEPT", -1) is not None else "")
        if dept not in ALLOWED_DEPT:
            continue

        no_ss = clean_text(row[idx.get("NO SS", 0)])
        if not no_ss:
            continue

        total += 1
        nama = clean_text(row[idx.get("PENCETUS IDE", -1)] if idx.get("PENCETUS IDE", -1) is not None else "")
        distrik_raw = clean_text(row[idx.get("DISTRIK", -1)] if idx.get("DISTRIK", -1) is not None else "KIDE")

        sql = text("""
            INSERT INTO tb_ss.ss_records 
              (no_ss, judul, nrp, nama, dept, divisi, distrik, tanggal_laporan,
               grade_ss, kualitas_ss, kategori_ss, reward_ss, status_karyawan,
               manfaat_financial, tanggal_menilai, dinilai_oleh, 
               current_status, source, latest_import_id)
            VALUES (:no_ss, :judul, :nrp, :nama, :dept, :divisi, :distrik, :tgl_lap,
                    :grade, :kualitas, :kategori, :reward, :status_kar, :financial,
                    :tgl_nilai, :dinilai, 'closed', 'closed', :import_id)
            ON CONFLICT (no_ss) DO UPDATE SET
              judul = EXCLUDED.judul, nrp = EXCLUDED.nrp, nama = EXCLUDED.nama,
              dept = EXCLUDED.dept, divisi = EXCLUDED.divisi, distrik = EXCLUDED.distrik,
              tanggal_laporan = EXCLUDED.tanggal_laporan, grade_ss = EXCLUDED.grade_ss,
              kualitas_ss = EXCLUDED.kualitas_ss, kategori_ss = EXCLUDED.kategori_ss,
              reward_ss = EXCLUDED.reward_ss, status_karyawan = EXCLUDED.status_karyawan,
              manfaat_financial = EXCLUDED.manfaat_financial,
              tanggal_menilai = EXCLUDED.tanggal_menilai,
              dinilai_oleh = EXCLUDED.dinilai_oleh,
              current_status = EXCLUDED.current_status,
              source = EXCLUDED.source, latest_import_id = EXCLUDED.latest_import_id
            RETURNING (xmax = 0) AS is_new
        """)

        res = await db.execute(sql, {
            "no_ss": no_ss,
            "judul": clean_text(row[idx.get("JUDUL", -1)] if idx.get("JUDUL", -1) is not None else ""),
            "nrp": clean_text(row[idx.get("NRP", -1)] if idx.get("NRP", -1) is not None else ""),
            "nama": nama,
            "dept": dept,
            "divisi": "",
            "distrik": distrik_raw or "KIDE",
            "tgl_lap": parse_date(row[idx.get("TANGGAL LAPORAN", -1)] if idx.get("TANGGAL LAPORAN", -1) is not None else None),
            "grade": clean_text(row[idx.get("GRADE SS", -1)] if idx.get("GRADE SS", -1) is not None else ""),
            "kualitas": clean_text(row[idx.get("KUALITAS SS", -1)] if idx.get("KUALITAS SS", -1) is not None else ""),
            "kategori": clean_text(row[idx.get("KATEGORI SS", -1)] if idx.get("KATEGORI SS", -1) is not None else ""),
            "reward": parse_num(row[idx.get("REWARD SS", -1)] if idx.get("REWARD SS", -1) is not None else 0) or 0,
            "status_kar": clean_text(row[idx.get("STATUS KARYAWAN", -1)] if idx.get("STATUS KARYAWAN", -1) is not None else ""),
            "financial": parse_num(row[idx.get("MANFAAT FINANCIAL", -1)] if idx.get("MANFAAT FINANCIAL", -1) is not None else None),
            "tgl_nilai": parse_date(row[idx.get("TANGGAL MENILAI", -1)] if idx.get("TANGGAL MENILAI", -1) is not None else None),
            "dinilai": clean_text(row[idx.get("DINILAI OLEH", -1)] if idx.get("DINILAI OLEH", -1) is not None else ""),
            "import_id": import_id,
        })
        if res.scalar():
            new_count += 1
        else:
            updated_count += 1

    return total, new_count, updated_count


async def _process_outstanding_records(db: AsyncSession, ws, import_id: int):
    """Process outstanding SS worksheet."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    new_count = 0
    updated_count = 0
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        dept = clean_text(row[idx.get("DEPT", -1)] if idx.get("DEPT", -1) is not None else "")
        if dept not in ALLOWED_DEPT:
            continue

        no_ss = clean_text(row[idx.get("NO SS", 0)])
        if not no_ss:
            continue

        total += 1
        distrik_raw = clean_text(row[idx.get("DISTRIK", -1)] if idx.get("DISTRIK", -1) is not None else "KIDE")

        sql = text("""
            INSERT INTO tb_ss.ss_records 
              (no_ss, judul, nrp, nama, dept, divisi, distrik, tanggal_laporan,
               grade_ss, kualitas_ss, kategori_ss, reward_ss, status_karyawan,
               manfaat_financial, tanggal_menilai, dinilai_oleh, 
               current_status, source, latest_import_id)
            VALUES (:no_ss, :judul, :nrp, :nama, :dept, :divisi, :distrik, :tgl_lap,
                    '', '', '', 0, '', NULL, NULL, '', :status, 'outstanding', :import_id)
            ON CONFLICT (no_ss) DO UPDATE SET
              judul = EXCLUDED.judul, nrp = EXCLUDED.nrp, nama = EXCLUDED.nama,
              dept = EXCLUDED.dept, divisi = EXCLUDED.divisi, distrik = EXCLUDED.distrik,
              tanggal_laporan = EXCLUDED.tanggal_laporan,
              current_status = EXCLUDED.current_status,
              source = EXCLUDED.source, latest_import_id = EXCLUDED.latest_import_id
            RETURNING (xmax = 0) AS is_new
        """)

        res = await db.execute(sql, {
            "no_ss": no_ss,
            "judul": clean_text(row[idx.get("JUDUL", -1)] if idx.get("JUDUL", -1) is not None else ""),
            "nrp": clean_text(row[idx.get("NRP", -1)] if idx.get("NRP", -1) is not None else ""),
            "nama": clean_text(row[idx.get("NAMA", -1)] if idx.get("NAMA", -1) is not None else ""),
            "dept": dept,
            "divisi": clean_text(row[idx.get("DIVISI", -1)] if idx.get("DIVISI", -1) is not None else ""),
            "distrik": distrik_raw or "KIDE",
            "tgl_lap": parse_date(row[idx.get("TANGGAL LAPORAN", -1)] if idx.get("TANGGAL LAPORAN", -1) is not None else None),
            "status": clean_text(row[idx.get("STATUS", -1)] if idx.get("STATUS", -1) is not None else ""),
            "import_id": import_id,
        })
        if res.scalar():
            new_count += 1
        else:
            updated_count += 1

    return total, new_count, updated_count


@router.post("/closed", response_model=ImportResult)
async def import_closed(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Upload Excel file closed SS. Filter: SPL2 + STYR only."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "File must be .xlsx or .xls")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    # Create import log
    log_res = await db.execute(
        text("INSERT INTO tb_ss.import_log (filename, source_type) VALUES (:fn, 'closed') RETURNING id"),
        {"fn": file.filename},
    )
    import_id = log_res.scalar()

    total, new_count, updated_count = await _process_closed_records(db, ws, import_id)

    await db.execute(
        text("UPDATE tb_ss.import_log SET total_rows=:t, new_records=:n, updated_records=:u WHERE id=:id"),
        {"t": total, "n": new_count, "u": updated_count, "id": import_id},
    )
    await db.commit()

    return ImportResult(
        filename=file.filename,
        source_type="closed",
        total_rows=total,
        new_records=new_count,
        updated_records=updated_count,
        import_id=import_id,
    )


@router.post("/outstanding", response_model=ImportResult)
async def import_outstanding(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Upload Excel file outstanding SS. Filter: SPL2 + STYR only."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "File must be .xlsx or .xls")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    log_res = await db.execute(
        text("INSERT INTO tb_ss.import_log (filename, source_type) VALUES (:fn, 'outstanding') RETURNING id"),
        {"fn": file.filename},
    )
    import_id = log_res.scalar()

    total, new_count, updated_count = await _process_outstanding_records(db, ws, import_id)

    await db.execute(
        text("UPDATE tb_ss.import_log SET total_rows=:t, new_records=:n, updated_records=:u WHERE id=:id"),
        {"t": total, "n": new_count, "u": updated_count, "id": import_id},
    )
    await db.commit()

    return ImportResult(
        filename=file.filename,
        source_type="outstanding",
        total_rows=total,
        new_records=new_count,
        updated_records=updated_count,
        import_id=import_id,
    )


@router.get("/history")
async def import_history(db: AsyncSession = Depends(get_db)):
    """Riwayat import."""
    res = await db.execute(
        text("SELECT * FROM tb_ss.import_log ORDER BY id DESC LIMIT 20")
    )
    return res.mappings().all()
